# -*- coding: utf-8 -*-
"""
Functions that can be used for cxcalc

For more info on all possible calculations and their variables:
https://www.chemaxon.com/marvin-archive/5_0_0/marvin/help/applications/cxcalc-calculations.html

@author: nune558
"""

from openpyxl import load_workbook
import time
import os
from subprocess import PIPE, Popen, check_output, STDOUT  # Help with command line
from molmass import Formula
from sys import stdout

#%% Functions


def _altcmdline(op, refine=False):
    process = Popen(['cmd', '/c', 'cxcalc.exe %s %s' % (op, inchi)], cwd=r'C:\Program Files\ChemAxon\MarvinSuite\bin\\', stdout=PIPE)
    if not refine:
        return process
    else:
        return process.communicate()[0].split('\n')[1].split('\t')[1].replace('\r', '')


def formula():
    f = _altcmdline('formula').communicate()[0]
    if f is None:
        return None
    f = f.split()[-1]
    return f

# Takes in compound and returns list of all possible tautomers at pH = 7.


def major_tautomer(allow_charge=False):
    taut = wait_timeout(_altcmdline('majortautomer -f inchi'), 120)
    if taut is None:
        return None
    return taut.communicate()[0].split('\n')[0][:-1]


def remove_charge(inchi):
    if 'q' in inchi:
        l = inchi.split('/')
        new = l[0]
        for i in range(1, len(l)):
            if 'q' not in l[i]:
                new += '/' + l[i]
        return new
    return inchi


def atomcount():
    return _cmdline('atomcount', refine=True)


def formal_charge():
    return _cmdline('formalcharge', refine=True)


def phys_charge():
    return _cmdline('formalcharge -H 7.4', refine=True)


def logP_ChemAxon():
    return _cmdline('logp', refine=True)


def pka_sa():
    return _cmdline('pka -t acidic -a 1 -d large', refine=True)


def pka_sb():
    return _cmdline('pka -t basic -b 1 -d large', refine=True)


def h_acc_count():
    return _cmdline('acceptorcount', refine=True)


def h_don_count():
    return _cmdline('donorcount', refine=True)


def polar_sa():
    return _cmdline('psa', refine=True)


def rbond_count():
    return _cmdline('rotatablebondcount', refine=True)


def refractivity():
    return _cmdline('refractivity', refine=True)


def polarizability():
    return _cmdline('tholepolarizability', refine=True)


def num_ionizable_sites():
    settings = '-a 10 -b 10 -i 0 -x 14 --considertautomerization True'
    l = _cmdline('pka ' + settings)
    l = l.split('\n')[1].split('\t')[1:21]
    return len([x for x in l if len(x) > 0])


def bondcount():
    return _cmdline('bondcount', refine=True)


def ringbondcount():
    return _cmdline('ringbondcount', refine=True)


def stereodoublebondcount():
    return _cmdline('stereodoublebondcount', refine=True)


def chainbondcount():
    return _cmdline('chainbondcount', refine=True)


def chiralcentercount():
    return _cmdline('chiralcentercount', refine=True)


def hararyindex():
    return _cmdline('hararyindex', refine=True)


def balabanindex():
    return _cmdline('balabanindex', refine=True)


def _cmdline(command, path):
    process = Popen(['cmd', '/c', command], cwd=path, stdout=PIPE, universal_newlines=True)
    for stdout_line in iter(process.stdout.readline, ''):
        yield stdout_line


def _cmdline_obabel(command):
    process = check_output(['cmd', '/c', command], cwd=r'C:\Program Files (x86)\OpenBabel-2.3.2\\')
    if len(process) == 0:
        return None
#    print(process)
    return process


def wait_timeout(proc, seconds):
    start = time.time()
    end = start + seconds
    interval = 10

    while True:
        result = proc.poll()
        if result is not None:
            return proc
        if time.time() >= end:
            proc.kill()
            return None
        time.sleep(interval)


def write_inchi(inchi, set_num):
    # Generate report
    r = open(r'C:\Users\nune558\Desktop\inchi%i.inchi' % set_num, 'w')
    r.write(inchi)
    r.close()


def desalt(inchi, set_num):

    # Desalt
    command = r'obabel -iinchi %s\inchi%i.inchi -r -osmi' % (r'C:\Users\nune558\Desktop', set_num)
    for t in _cmdline_obabel(command):
        smi = _cmdline_obabel(command)

        if smi is None:
            return None

        smi = smi[:-3].replace('\"', '')
        command = r'obabel.exe -:%s -oinchi' % smi

        for t in _cmdline_obabel(command):
            return t


def smi2inchi(smi):

    # Desalt
    command = r'obabel -:%s -oinchi' % (smi)
    temp = _cmdline_obabel(command)
#    print(command)
    if temp is None:
        return None
    for t in temp:
        smi = _cmdline_obabel(command)

        if smi is None:
            return None

        return smi

#%% Temp Main


path = r'C:\Users\nune558\Google Drive\Jamie Nunez PNNL\EPA\Final Spreadsheets/'
fname = 'ToxCast4000_Final_ISICLE_Key_temp'
wb = load_workbook(path + fname + '.xlsx')
sheet = wb.worksheets[0]

set_num = 0
og_start_row = 4239
stop_row = sheet.max_row
#fname = path + 'tautomer_tempfile_%i.txt' % og_start_row
# if True: # Start new run
#    start_row = og_start_row
#    f = open(fname, 'w')
#    f.write('ID\tMajor Tautomer\tFormula\tMass\tpKa\tlogP\tHInd\tBInd\tRB%\n')
# else: # Continue run after error
#    start_row = 24
# with open(fname) as f:
##        start_row = len(f.readlines()) + og_start_row
#    f = open(fname, 'a')
#    f.write('\n')
#
# print 'min: %s, max: %s' % (og_start_row, stop_row)

t = time.time()
id_col = 'B'
inchi_col = 'F'

for i in range(og_start_row, stop_row + 1):

    # Update console
    #    if i % 1000 == 0:
    #        print i
    #        print 'Time for last 1000: %.2f' % (time.time() - t)
    #        stdout.flush()
    #        f.flush()
    #        os.fsync(f)
    #        t = time.time()

    #    print i
    #    stdout.flush()
    #    f.flush()
    #    os.fsync(f)
    #    t = time.time()

    #    s = sheet[id_col + str(i)].value
    if True:  # sheet['E' + str(i)].value < 1200:
        inchi = sheet[inchi_col + str(i)].value
#        sheet['C' + str(i)] = smi2inchi(inchi)

        inchi = remove_charge(inchi)
        write_inchi(inchi, set_num)
        inchi = desalt(inchi, set_num)
#        write_inchi(inchi, set_num)
#        if inchi is not None:
#            stdout.flush()
#            inchi = major_tautomer(inchi)

        # Get properties
        if inchi is not None and len(inchi) > 0:

            #            inchi = inchi.split('\n')[0]
            #            s = inchi
            #            s += ('\t' + inchi)
            if '/p' in inchi or '/i' in inchi:
                form = formula()
            else:
                form = inchi.split('/')[1]
#            s += ('\t' + form)
#            s += ('\t' + str(Formula(form).isotope.mass))

            sheet['G' + str(i)] = inchi
            sheet['H' + str(i)] = form
            sheet['I' + str(i)] = Formula(form).isotope.mass
#            s += ('\t' + str(pka_sa()))
#            s += ('\t' + str(logP_ChemAxon()))
#            s += ('\t' + str(hararyindex()))
#            s += ('\t' + str(balabanindex()))
#            bc = float(bondcount())
#            if bc != 0:
#                s += ('\t' + str(float(ringbondcount()) / bc))
#            else:
#                 s += ('\t')

#    f.write(s + '\n')
# f.close()


wb.save(path + fname + '_temp.xlsx')
